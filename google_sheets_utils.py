# File: google_sheets_utils.py (Improved Version with add_new_product_rows)

import os
import gspread # Requires: pip install gspread
import logging
# Requires: pip install oauth2client
from oauth2client.service_account import ServiceAccountCredentials
import time
from threading import Lock
import requests # Requires: pip install requests
import json # Needed for header check in move_done
import random # Needed for jitter
import gspread
from openpyxl.utils import get_column_letter

# --- Type Hinting Imports ---
from typing import List, Dict, Any, Tuple, Optional, Union, Set

# --- Setup logger ---
# Use standard logger name, format allows easy filtering
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - [%(name)s] - %(message)s')
logger = logging.getLogger(__name__)

# --- Configuration ---
# Load from environment variables, provide defaults where sensible
GOOGLE_CREDENTIALS_FILE = os.getenv("GOOGLE_CREDENTIALS_FILE", "credentials.json")
GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID") # CRITICAL - Must be set in .env
SHEET1_NAME = os.getenv("STATUS_SHEET_NAME", "Sheet1") # Main processing sheet
SHEET2_NAME = os.getenv("ARCHIVE_SHEET_NAME", "Sheet2") # Archive sheet

MAX_RETRIES = 50     # Max retries for Google API calls
RETRY_DELAY = 5.0   # Initial delay in seconds for retries

# --- Column Mapping for Multi-Store Status Updates ---
# !!! IMPORTANT: Verify these column letters match your ACTUAL Sheet1 layout !!!
STORE_COLUMN_MAP = {
    "store_es": {"status": 4, "gid": 5, "title": 6},  # D=4, E=5, F=6
    "store_dk": {"status": 7, "gid": 8, "title": 9},  # G=6, H=7, I=8 (0-based)
}   
# --- Helper Functions ---
_gspread_client = None
_client_lock = Lock()

# At the top of your google_sheets_utils.py
STORE_COLUMN_MAP = {
    "store_es": {"status": 4, "gid": 5, "title": 6},  # Example: Status in D, GID in E, Title in F
    "store_dk": {"status": 7, "gid": 8, "title": 9},  # Example: Status in G, GID in H, Title in I
    # Add other stores from your SHOPIFY_STORES_CONFIG
}

def clean_header_key(header: str) -> str:
    """Removes leading/trailing spaces and normalizes internal whitespace."""
    if not isinstance(header, str):
        return str(header) # Return string representation if not a string
    # Remove leading/trailing whitespace
    cleaned = header.strip()
    # Optional: Replace multiple internal spaces with a single space
    # cleaned = ' '.join(cleaned.split())
    # Optional: Remove specific non-printing chars if needed (example: remove newline)
    # cleaned = cleaned.replace('\n', '').replace('\r', '')
    return cleaned

def _get_gspread_client() -> Optional[gspread.Client]:
    """Authenticates and returns an authorized gspread client (cached). Thread-safe."""
    global _gspread_client
    # Quick check without lock first
    if _gspread_client:
        return _gspread_client

    with _client_lock:
        # Double-check after acquiring lock
        if _gspread_client is None:
            scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
            creds_path = GOOGLE_CREDENTIALS_FILE
            if not GOOGLE_SHEET_ID:
                 logger.critical("❌ CRITICAL: GOOGLE_SHEET_ID environment variable not set.")
                 return None # Cannot proceed without Sheet ID
            if not os.path.exists(creds_path):
                logger.critical(f"❌ Google Credentials file not found at: {creds_path}. Check GOOGLE_CREDENTIALS_FILE env var.")
                return None
            try:
                logger.debug(f"Attempting to authorize Google Sheets client using: {creds_path}")
                creds = ServiceAccountCredentials.from_json_keyfile_name(creds_path, scope)
                client = gspread.authorize(creds)
                logger.info(f"🔐 Google Sheets client authorized successfully using: {creds_path}")
                _gspread_client = client
            except Exception as e:
                logger.exception(f"❌ Failed to authorize Google Sheets client: {e}")
                _gspread_client = None # Ensure it stays None on failure
    return _gspread_client

def _get_worksheet(worksheet_name: str) -> Optional[gspread.Worksheet]:
    """Gets a specific worksheet object using the cached client."""
    client = _get_gspread_client()
    if not client:
        logger.error("❌ Cannot get worksheet, client not authorized.")
        return None
    # GOOGLE_SHEET_ID check happens in client auth now

    try:
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
        sheet = spreadsheet.worksheet(worksheet_name)
        logger.debug(f"Accessed worksheet: '{worksheet_name}' in Sheet ID: {GOOGLE_SHEET_ID}")
        return sheet
    except gspread.WorksheetNotFound:
        logger.error(f"❌ Worksheet '{worksheet_name}' not found in Google Sheet ID {GOOGLE_SHEET_ID}.")
    except gspread.exceptions.APIError as e:
        # Log specific API errors (like permission denied, quota exceeded)
        response_text = getattr(e.response, 'text', str(e)) # Safely get response text
        status_code = getattr(e.response, 'status_code', 'N/A')
        logger.error(f"❌ API Error opening worksheet '{worksheet_name}' (SheetID: {GOOGLE_SHEET_ID}): Status {status_code}, Details: {response_text[:500]}...") # Limit error text length
    except Exception as e:
        logger.exception(f"❌ Failed to open worksheet '{worksheet_name}' (SheetID: {GOOGLE_SHEET_ID}): {e}")
    return None

def _retry_gspread_operation(operation: callable, *args, **kwargs) -> Any:
    """Wrapper to retry gspread operations with exponential backoff and jitter."""
    retries = 0
    last_exception = None
    while retries < MAX_RETRIES:
        try:
            # Attempt the operation (e.g., sheet.get_all_values(), sheet.append_rows())
            return operation(*args, **kwargs)
        except requests.exceptions.ConnectionError as conn_err:
            # Network-level errors
            last_exception = conn_err; retries += 1; error_type = "ConnectionError"
        except gspread.exceptions.APIError as api_err:
            # Errors specifically from the Google Sheets API
            last_exception = api_err; error_type = "APIError"
            # Check if error is retryable (Rate Limit 429, Server Errors 5xx)
            status_code = getattr(api_err.response, 'status_code', 0)
            if status_code == 429 or status_code >= 500:
                retries += 1 # Retry these errors
            else:
                # Do not retry client errors like 400, 401, 403, 404
                response_text = getattr(api_err.response, 'text', str(api_err))
                logger.error(f"❌ Non-retryable Google Sheets API Error during {operation.__name__}: Status {status_code}, Details: {response_text[:500]}...")
                raise # Re-raise the exception to stop the process
        except Exception as e:
             # Catch any other unexpected exceptions during the gspread call
             last_exception = e; error_type = "Unexpected Error"
             logger.error(f"❌ {error_type} during {operation.__name__}: {e}", exc_info=True)
             raise # Re-raise unexpected errors immediately

        # If a retry is needed and possible:
        if retries >= MAX_RETRIES:
            logger.error(f"🚫 Max retries ({MAX_RETRIES}) exceeded for {operation.__name__} after {error_type}: {last_exception}")
            raise last_exception # Re-raise the last captured exception

        # Calculate wait time with exponential backoff and jitter
        wait = RETRY_DELAY * (2 ** (retries - 1)) * (1 + random.random())
        wait = min(wait, 30.0) # Cap wait time
        status_code_info = f" (Status: {status_code})" if error_type == "APIError" else ""
        logger.warning(f"🔁 {error_type}{status_code_info} during {operation.__name__}. Retrying in {wait:.2f}s (Attempt {retries}/{MAX_RETRIES})...")
        time.sleep(wait)

    # Fallback raise if loop structure somehow allows exit without raising
    raise last_exception if last_exception else Exception(f"Max retries exceeded for {operation.__name__}")

# --- Public Utility Functions ---

def ensure_sheet_headers(sheet_name: str) -> Tuple[Optional[List[str]], bool]:
    """
    Checks Sheet headers and updates them if necessary to match the FIXED standard layout:
    A=Product ID, B=Product Title, C=Sales Count,
    D=Status ES, E=Cloned GID ES, F=Cloned Title ES,
    G=Status DK, H=Cloned GID DK, I=Cloned Title DK
    Returns (final_headers_list or None, success_boolean).
    """
    logger.info(f"Checking/Ensuring standard fixed headers for sheet '{sheet_name}'...")

    sheet = None
    current_header: Optional[List[str]] = None # To store headers read from sheet

    try:
        # Get the worksheet object
        sheet = _get_worksheet(sheet_name)
        if not sheet:
            logger.error(f"Failed to get worksheet '{sheet_name}' in ensure_sheet_headers.")
            return None, False
    except Exception as e:
         logger.exception(f"Error getting worksheet '{sheet_name}': {e}")
         return None, False

    # --- Define the FIXED expected header structure ---
    expected_headers = [
        "Product ID",           # Col A
        "Product Title",        # Col B
        "Sales Count",          # Col C
        "Status ES",            # Col D (Spanish Status)
        "Cloned GID ES",        # Col E (Spanish GID)
        "Cloned Title ES",      # Col F (Spanish Title)
        "Status DK",            # Col G (Danish Status)
        "Cloned GID DK",        # Col H (Danish GID)
        "Cloned Title DK"       # Col I (Danish Title)
    ]
    # --- End fixed definition ---

    logger.debug(f"Expected Fixed Headers: {expected_headers}")

    try:
        # Attempt to read the current header row
        try:
            # --- FIX 1: Use openpyxl utility ---
            # Calculate column letter for fetching (e.g., N for 9 + 5 = 14)
            fetch_range_end_col_letter = get_column_letter(len(expected_headers) + 5)
            fetch_range = f"A1:{fetch_range_end_col_letter}1"
            header_values = _retry_gspread_operation(sheet.get_values, fetch_range)

            if header_values and isinstance(header_values, list) and len(header_values) > 0:
                 current_header = header_values[0]
                 while current_header and current_header[-1] == '':
                     current_header.pop() # Trim empty cells from end
            else:
                 logger.warning(f"Could not read header row from '{sheet_name}' or it was empty.")
                 current_header = []
        except Exception as read_err:
            logger.error(f"Error reading header row from '{sheet_name}': {read_err}")
            current_header = [] # Assume empty on error

        logger.debug(f"Current Headers read (or empty): {current_header}")

        # Compare with the fixed list
        if current_header == expected_headers:
            logger.info(f"✅ Headers in '{sheet_name}' match the expected fixed layout.")
            return expected_headers, True # Success, headers are correct

        # --- Headers need update ---
        logger.warning(f"Headers mismatch/missing from expected fixed layout. Attempting update in '{sheet_name}'...")

        # --- FIX 2: Use openpyxl utility ---
        # Calculate end column letter (e.g., I for 9 columns)
        header_range_end_col_letter = get_column_letter(len(expected_headers))
        header_range = f"A1:{header_range_end_col_letter}1" # e.g., "A1:I1"
        logger.debug(f"Updating header range: {header_range}")

        # --- Optional: Clear extra columns beyond the expected range ---
        if current_header is not None and len(current_header) > len(expected_headers):
            try:
                clear_start_col_num = len(expected_headers) + 1
                # --- FIX 3: Use openpyxl utility ---
                clear_start_col_letter = get_column_letter(clear_start_col_num) # e.g., J
                # Clear from J1 potentially up to Z1 (adjust Z if more cols needed)
                clear_range = f"{clear_start_col_letter}1:Z1"
                logger.debug(f"Clearing any extra header cells in range: {clear_range}")
                _retry_gspread_operation(sheet.batch_clear, [clear_range])
            except Exception as clear_err:
                 logger.warning(f"⚠️ Failed to clear potential extra header cells: {clear_err}. Continuing update...")
        # --- End Optional Clearing ---

        # Update the calculated range (e.g., A1:I1) with the fixed expected headers
        _retry_gspread_operation(sheet.update, header_range, [expected_headers], value_input_option="USER_ENTERED")
        logger.info(f"✅ Headers updated in '{sheet_name}' to match fixed layout.")
        return expected_headers, True # Return the headers we just wrote

    except gspread.exceptions.APIError as gae:
         logger.error(f"❌ gspread API error during header check/update for '{sheet_name}': {gae}")
         return current_header, False # Return read headers (if any), flag failure
    except Exception as e:
        # Catch any other unexpected errors during the process
        logger.exception(f"❌ Unexpected error checking/updating headers for '{sheet_name}': {e}")
        return current_header, False # Return read headers (if any), flag failure

# --- NEW FUNCTION TO ADD ROWS ---
def add_new_product_rows(rows_to_add: List[List[Any]], sheet_name: str = SHEET1_NAME) -> bool:
    """
    Appends multiple rows of data to the specified sheet.

    Args:
        rows_to_add (List[List[Any]]): A list of lists, where each inner list
                                       represents a row and contains the cell values
                                       in the correct column order.
        sheet_name (str): The name of the target worksheet.

    Returns:
        bool: True if rows were appended successfully, False otherwise.
    """
    if not rows_to_add:
        logger.info(f"No new rows provided to add to '{sheet_name}'.")
        return True # Nothing to do is considered success

    logger.info(f"Attempting to append {len(rows_to_add)} new rows to sheet '{sheet_name}'...")
    sheet = _get_worksheet(sheet_name)
    if not sheet:
        logger.error(f"Cannot append rows, failed to get worksheet '{sheet_name}'.")
        return False

    try:
        # Use the append_rows method with retry logic
        _retry_gspread_operation(
            sheet.append_rows,
            values=rows_to_add,
            value_input_option="USER_ENTERED", # Interprets values (e.g., numbers)
            insert_data_option="INSERT_ROWS" # Appends after the last row with content
        )
        logger.info(f"✅ Successfully appended {len(rows_to_add)} rows to '{sheet_name}'.")
        return True
    except Exception as e:
        # Error likely logged by retry wrapper, add context here
        logger.exception(f"❌ Failed to append rows to '{sheet_name}': {e}")
        return False
# --- END NEW FUNCTION ---


def find_row_index_by_id(sheet: gspread.Worksheet, product_id: Union[str, int], id_column_index: int = 1) -> Optional[int]:
     """Finds the row index (1-based) for a given ID in a specific column using retry."""
     if not sheet or not product_id: return None
     try:
         # Find expects a string query
         cell = _retry_gspread_operation(sheet.find, str(product_id), in_column=id_column_index)
         if cell:
              logger.debug(f"Found product_id '{product_id}' in row {cell.row}")
              return cell.row
         else:
              logger.debug(f"Product_id '{product_id}' not found in column {id_column_index}.")
              return None
     except Exception as e:
         # Error likely logged by retry wrapper
         logger.error(f"Error finding row for ID {product_id} in col {id_column_index}: {e}")
         return None

def update_export_status_for_store(
    original_product_id: Union[str, int], 
    target_store_value: str, # This should be the store key like "store_es"
    status_value: str,       # Parameter name is now status_value
    cloned_gid: Optional[str], 
    cloned_title: Optional[str],
    sheet_name: str = SHEET1_NAME 
) -> bool:
    """
    Updates Status, GID, and Title using column INDICES defined in STORE_COLUMN_MAP.
    Converts indices to letters before creating A1 range notation for update.
    """
    original_product_id_str = str(original_product_id)
    if not original_product_id_str or not target_store_value:
        logger.warning("⚠️ update_export_status_for_store: Missing product_id or target_store_value.")
        return False

    if 'STORE_COLUMN_MAP' not in globals() or not isinstance(STORE_COLUMN_MAP, dict) or not STORE_COLUMN_MAP:
        logger.error("❌ STORE_COLUMN_MAP is not defined, not a dictionary, or is empty in google_sheets_utils.")
        return False
        
    column_set = STORE_COLUMN_MAP.get(target_store_value)
    if not column_set:
        logger.error(f"❌ No column mapping found for target store '{target_store_value}' in STORE_COLUMN_MAP.")
        return False

    status_col_idx = column_set.get('status')
    gid_col_idx = column_set.get('gid')
    title_col_idx = column_set.get('title')

    if not (isinstance(status_col_idx, int) and status_col_idx > 0 and
            isinstance(gid_col_idx, int) and gid_col_idx > 0 and
            isinstance(title_col_idx, int) and title_col_idx > 0):
         logger.error(f"❌ Invalid column index in STORE_COLUMN_MAP for '{target_store_value}'. Found: {column_set}")
         return False

    log_title_snip = str(cloned_title)[:50] + '...' if cloned_title else '(empty)'
    # Corrected to use status_value in the log message
    logger.info(f"Sheet Update Prep for Orig ID {original_product_id_str} / Store '{target_store_value}': "+
                f"Status='{status_value}'(ColIdx:{status_col_idx}), GID='{cloned_gid or ''}'(ColIdx:{gid_col_idx}), Title='{log_title_snip}'(ColIdx:{title_col_idx})")

    try:
        sheet = _get_worksheet(sheet_name)
        if not sheet: 
            logger.error(f"Failed to get worksheet: {sheet_name}")
            return False
        
        id_column_index_for_find = 1 # Defaulting to column A for Product ID lookup
        
        row_index = find_row_index_by_id(sheet, original_product_id_str, id_column_index=id_column_index_for_find) 

        if row_index and isinstance(row_index, int) and row_index > 0:
            try:
                status_col_letter = get_column_letter(status_col_idx)
                gid_col_letter = get_column_letter(gid_col_idx)
                title_col_letter = get_column_letter(title_col_idx)
            except Exception as conversion_err:
                 logger.error(f"❌ Failed to convert column indices to letters: {conversion_err}")
                 return False

            status_range = f"{status_col_letter}{row_index}"
            gid_range = f"{gid_col_letter}{row_index}"
            title_range = f"{title_col_letter}{row_index}"

            logger.debug(f"Calculated ranges: Status='{status_range}', GID='{gid_range}', Title='{title_range}'")

            # Corrected to use status_value when preparing update data
            update_data = [
                {"range": status_range, "values": [[str(status_value or '')]]},
                {"range": gid_range,    "values": [[str(cloned_gid or '')]]},
                {"range": title_range,  "values": [[str(cloned_title or '')]]}
            ]
            logger.debug(f"Attempting batch update for row {row_index} with data: {update_data}")
            _retry_gspread_operation(sheet.batch_update, update_data, value_input_option="USER_ENTERED")
            logger.info(f"✅ Updated Sheet Row {row_index} for Store '{target_store_value}'.")
            return True
        else:
            logger.warning(f"⚠️ Could not find valid row index for Product ID '{original_product_id_str}' in '{sheet_name}' (Col {id_column_index_for_find}).")
            return False
    except Exception as e:
        logger.exception(f"❌ Error in update_export_status_for_store for ID {original_product_id_str} / Store '{target_store_value}': {e}")
        return False

# In google_sheets_utils.py
# Replace the existing get_sheet_data_by_header function with this:

def get_sheet_data_by_header(sheet_name: str = SHEET1_NAME) -> Tuple[Optional[List[str]], Optional[List[Dict[str, str]]]]:
    """
    Fetches all data from a sheet using the first row as headers.
    Uses get_all_values() and manually creates the list of dictionaries.
    Returns the header row as a list and the data rows as a list of dictionaries.
    Returns (None, None) on failure.
    """
    logger.info(f"Fetching all data (using get_all_values) with headers from sheet '{sheet_name}'...")
    sheet = _get_worksheet(sheet_name)
    if not sheet:
        logger.error(f"Cannot fetch data, failed to get worksheet '{sheet_name}'.")
        return None, None

    try:
        # Fetch ALL cell values as a list of lists using retry
        all_values = _retry_gspread_operation(sheet.get_all_values)

        if not all_values:
            logger.warning(f"Sheet '{sheet_name}' appears to be empty or could not be read via get_all_values.")
            return None, None # Return None if sheet is empty or read failed

        # First row is the header
        header = all_values[0]
        # Data rows are the rest
        data_rows = all_values[1:]

        if not header:
             logger.error(f"Sheet '{sheet_name}' has data rows but the header row is empty.")
             return None, None # Cannot proceed without headers

        if not data_rows:
            logger.info(f"Sheet '{sheet_name}' has header but no data rows.")
            return header, [] # Return header and empty list

        # Manually convert data rows to list of dictionaries
        list_of_dicts = []
        header_len = len(header)
        for row_idx, row_values in enumerate(data_rows, start=2): # Start from row 2 for logging
            # Pad row with empty strings if shorter than header
            padded_row = row_values + [''] * (header_len - len(row_values))
            # Create dict using header as keys, ensuring values are strings
            row_dict = {str(h): str(padded_row[i]) for i, h in enumerate(header)}
            list_of_dicts.append(row_dict)

        logger.info(f"Successfully fetched header ({len(header)} columns) and {len(list_of_dicts)} data rows from '{sheet_name}' (using get_all_values).")
        return header, list_of_dicts

    except Exception as e:
        # Catch errors during get_all_values or dict conversion
        logger.exception(f"❌ Failed to get/process sheet data using get_all_values for '{sheet_name}': {e}")
        return None, None
    
# In google_sheets_utils.py
# Add necessary imports if missing: from typing import Optional, Set
# Make sure ARCHIVE_SHEET_NAME is defined (e.g., ARCHIVE_SHEET_NAME = "Sheet2")

ARCHIVE_SHEET_NAME = "Sheet2"

# In google_sheets_utils.py
# Make sure ARCHIVE_SHEET_NAME is defined (e.g., from os.getenv)
# Make sure logger, _get_worksheet, _retry_gspread_operation are defined

def get_archived_product_ids(sheet_name: str = ARCHIVE_SHEET_NAME) -> Optional[Set[str]]: # Needs "Set" and "Optional" from typing
    """
    Fetches all Product IDs from the first column of the specified archive sheet.
    Assumes the first column contains Product IDs and the first row is a header.
    Returns a set of Product IDs (as strings) or None on failure.
    """
    logger.info(f"Fetching archived Product IDs from column A of sheet '{sheet_name}'...")
    sheet = _get_worksheet(sheet_name)
    if not sheet:
        logger.error(f"Cannot fetch archived IDs, failed to get worksheet '{sheet_name}'.")
        return None

    try:
        # Fetch all values from the first column (A) using retry
        id_list = _retry_gspread_operation(sheet.col_values, 1) # col_values uses 1-based index

        if not id_list or len(id_list) <= 1: # Check if list is empty or only contains header
            logger.info(f"No archived Product IDs found (or only header) in column A of '{sheet_name}'.")
            return set() # Uses built-in set() constructor

        # Assume first row is header, skip it [1:]. Convert rest to strings and filter empties.
        # Uses set comprehension {}
        archived_ids = {str(pid).strip() for pid in id_list[1:] if str(pid).strip()}

        logger.info(f"Successfully fetched {len(archived_ids)} unique archived Product IDs from '{sheet_name}'.")
        return archived_ids

    except Exception as e:
        logger.exception(f"❌ Failed to get archived Product IDs from '{sheet_name}': {e}")
        return None
    
def move_done_to_sheet2():
    """
    Moves completed rows (based on STORE_COLUMN_MAP statuses like DONE_*, TRANSLATED, APPROVED)
    from Sheet1 to Sheet2 and deletes them from Sheet1 using batch operations.
    """
    logger.info(f"Checking for completed rows to move from '{SHEET1_NAME}' to '{SHEET2_NAME}'...")
    sheet1 = _get_worksheet(SHEET1_NAME)
    sheet2 = _get_worksheet(SHEET2_NAME)
    if not sheet1 or not sheet2:
        logger.error(f"Could not access '{SHEET1_NAME}' or '{SHEET2_NAME}'. Aborting move.")
        return

    try:
        # Fetch header and data using the reliable helper
        header, all_data = get_sheet_data_by_header(SHEET1_NAME)
        if header is None or all_data is None:
            logger.error(f"Failed to fetch data from '{SHEET1_NAME}'. Aborting move.")
            return
        if not all_data:
            logger.info(f"📭 No data rows found in '{SHEET1_NAME}' to process for moving.")
            return

        # Dynamically determine which status columns to check based on the header AND the map
        actual_status_columns = []
        for store_key, cols in STORE_COLUMN_MAP.items():
             status_col_header = f"Status {store_key.split('_')[-1].upper()}"
             if 'status' in cols and status_col_header in header: # Check if map defined AND exists in sheet
                 actual_status_columns.append(status_col_header)
        # Optionally add base 'Status' column if it exists
        if "Status" in header and "Status" not in actual_status_columns:
             actual_status_columns.append("Status")

        if not actual_status_columns:
            logger.error(f"❌ Cannot find any expected Status columns based on STORE_COLUMN_MAP in header of '{SHEET1_NAME}'. Headers: {header}")
            return
        logger.debug(f"Status columns being checked for completion: {actual_status_columns}")

        # Define statuses that trigger a move
        statuses_to_move = {"APPROVED", "TRANSLATED"}
        # Dynamically add DONE_STORE_XYZ statuses based on map keys
        for store_key in STORE_COLUMN_MAP.keys():
            statuses_to_move.add(f"DONE_{store_key.upper()}") # e.g., DONE_STORE_ES
        logger.info(f"Statuses triggering move to Sheet2: {statuses_to_move}")

        rows_to_move_data = []
        rows_to_delete_requests = [] # Build requests for batchUpdate deleteDimension
        sheet1_gid = sheet1.id # Get the numeric GID of Sheet1

        # Iterate backwards through the indices to handle deletions correctly
        for idx in range(len(all_data) - 1, -1, -1):
            row_dict = all_data[idx]
            current_sheet_row_index = idx + 2 # Sheet index is 1-based, +1 for header

            # Check all relevant status columns for a "move" status
            should_move = False
            status_found = ""
            for status_col_name in actual_status_columns:
                status_val = str(row_dict.get(status_col_name, "")).strip().upper()
                if status_val in statuses_to_move:
                    should_move = True
                    status_found = status_val
                    break # Found a status in this row, no need to check others

            if should_move:
                logger.debug(f"Marking sheet row {current_sheet_row_index} for moving (Status: '{status_found}').")
                # Convert dict back to list in header order for appending
                row_values = [row_dict.get(h, "") for h in header]
                rows_to_move_data.append(row_values)
                # Add a delete request (API uses 0-based index)
                rows_to_delete_requests.append({
                    "deleteDimension": {
                        "range": {
                            "sheetId": sheet1_gid,
                            "dimension": "ROWS",
                            "startIndex": current_sheet_row_index - 1, # 0-based start index
                            "endIndex": current_sheet_row_index        # Exclusive end index
                        }
                    }
                })

        # --- Perform Sheet Updates (if any rows marked) ---
        if rows_to_move_data:
            # Append data to Sheet2 (reverse collected data back to original order for append)
            logger.info(f"Appending {len(rows_to_move_data)} rows to '{SHEET2_NAME}'...")
            _retry_gspread_operation(sheet2.append_rows, rows_to_move_data[::-1], value_input_option="USER_ENTERED", insert_data_option="INSERT_ROWS")

            # Batch delete rows from Sheet1 (API requires requests sorted descending by index)
            logger.info(f"Batch deleting {len(rows_to_delete_requests)} rows from '{SHEET1_NAME}'...")
            sorted_delete_requests = sorted(rows_to_delete_requests, key=lambda x: x['deleteDimension']['range']['startIndex'], reverse=True)
            delete_body = {'requests': sorted_delete_requests}
            _retry_gspread_operation(sheet1.batch_update, delete_body)

            logger.info(f"✅ Successfully moved and deleted {len(rows_to_move_data)} completed rows.")
        else:
            logger.info("📭 No completed rows found in Sheet1 to move.")

    except Exception as e:
        # Error logged by retry wrapper or here
        logger.exception(f"❌ Failed during move_done_to_sheet2 process: {e}")

# Add this function to google_sheets_utils.py
# Make sure STORE_COLUMN_MAP, SHEET1_NAME, ARCHIVE_SHEET_NAME are accessible
# Ensure logger, _get_worksheet, _retry_gspread_operation, get_sheet_data_by_header exist

def move_fully_done_to_sheet2():
    """
    Moves rows where ALL configured stores have a 'DONE_STORE_*' status
    from Sheet1 to Sheet2 and deletes them from Sheet1 using worksheet.delete_rows().
    Returns True if processing completed (whether rows were moved or not), False on error.
    """
    logger.info(f"Checking for fully completed rows to move from '{SHEET1_NAME}' to '{ARCHIVE_SHEET_NAME}'...")
    sheet1 = _get_worksheet(SHEET1_NAME)
    sheet2 = _get_worksheet(ARCHIVE_SHEET_NAME)
    if not sheet1 or not sheet2:
        logger.error(f"Could not access '{SHEET1_NAME}' or '{ARCHIVE_SHEET_NAME}'. Aborting move.")
        return False # Indicate failure to access sheets

    try:
        # Fetch header and data using the reliable helper
        header, all_data = get_sheet_data_by_header(SHEET1_NAME) # Assumes this reads fresh data
        if header is None or all_data is None:
            logger.error(f"Failed to fetch data from '{SHEET1_NAME}'. Aborting move.")
            return False
        if not all_data:
            logger.info(f"📭 No data rows found in '{SHEET1_NAME}' to process for moving.")
            return True # No rows to move is not an error

        # Get configured store keys and required headers
        # Ensure STORE_COLUMN_MAP is accessible here
        store_keys = list(STORE_COLUMN_MAP.keys())
        required_status_headers = []
        all_headers_present = True
        for store_key in store_keys:
             # Get configured header name (ensure SHOPIFY_STORES_CONFIG is parsed before this)
             # This part requires access to the parsed config or assumes STORE_COLUMN_MAP holds headers directly
             # For simplicity, let's assume STORE_COLUMN_MAP holds header names now based on previous fixes
             # If STORE_COLUMN_MAP holds indices, need to get header name from the sheet1_header list
             status_header = STORE_COLUMN_MAP.get(store_key, {}).get('sheet_status_col_header') # Assumes map keys match store_config keys
             if not status_header: # Fallback calculation if config key missing
                  store_suffix = store_key.split('_')[-1].upper()
                  status_header = f"Status {store_suffix}"
                  logger.warning(f"Using calculated status header '{status_header}' for store '{store_key}' in move function.")

             if status_header not in header:
                  logger.error(f"Required status header '{status_header}' for store '{store_key}' not found in sheet header: {header}. Aborting move.")
                  all_headers_present = False
                  break
             required_status_headers.append(status_header)

        if not all_headers_present:
             return False # Cannot proceed if expected headers are missing

        logger.debug(f"Checking completion across status columns: {required_status_headers}")

        rows_to_move_data = []
        row_indices_to_delete = [] # Store 1-based indices

        # Iterate backwards through the indices (0-based for list `all_data`)
        for idx in range(len(all_data) - 1, -1, -1):
            row_dict = all_data[idx]
            current_sheet_row_index = idx + 2 # Sheet index is 1-based, +1 for header

            # --- Check if ALL required stores are DONE ---
            all_stores_done = True
            if not store_keys: # Handle empty config case
                 all_stores_done = False
            for i, store_key in enumerate(store_keys):
                status_col_name = required_status_headers[i] # Get the correct header name
                # Construct the specific DONE status string for this store
                expected_done_status = f"DONE_{store_key.upper()}" # e.g., DONE_STORE_ES
                status_val = str(row_dict.get(status_col_name, "")).strip().upper()

                if status_val != expected_done_status:
                    all_stores_done = False
                    break # No need to check other stores for this row
            # --- End Check ---

            if all_stores_done:
                logger.info(f"Marking sheet row {current_sheet_row_index} for moving (All configured stores DONE). PID: {row_dict.get('Product ID', 'N/A')}")
                # Convert dict back to list in header order for appending
                row_values = [row_dict.get(h, "") for h in header]
                rows_to_move_data.append(row_values)
                # Store the 1-based sheet row index for deletion
                row_indices_to_delete.append(current_sheet_row_index)

        # --- Perform Sheet Updates ---
        if rows_to_move_data:
            logger.info(f"Moving {len(rows_to_move_data)} fully completed rows...")
            # Append data to Sheet2 (reverse collected data back to original order)
            logger.info(f"Appending {len(rows_to_move_data)} rows to '{SHEET2_NAME}'...")
            _retry_gspread_operation(sheet2.append_rows, rows_to_move_data[::-1], value_input_option="USER_ENTERED", insert_data_option="INSERT_ROWS")

            # --- Delete rows from Sheet1 using delete_rows() ---
            # Indices are already collected in descending order because we iterated backwards
            logger.info(f"Deleting {len(row_indices_to_delete)} rows from '{SHEET1_NAME}'...")
            deleted_count = 0
            for row_index_to_delete in row_indices_to_delete:
                 try:
                     logger.debug(f"Attempting to delete row {row_index_to_delete} from {SHEET1_NAME}...")
                     # Use the correct gspread method, wrapped in retry
                     _retry_gspread_operation(sheet1.delete_rows, row_index_to_delete)
                     deleted_count += 1
                     logger.debug(f"Successfully deleted row {row_index_to_delete}.")
                     time.sleep(0.3) # Add a small delay between deletes to be kind to API
                 except Exception as delete_err:
                     # Log error but continue trying to delete others
                     # Error should be caught by _retry_gspread_operation first if it's APIError/ConnError
                     logger.error(f"Failed to delete row {row_index_to_delete} from '{SHEET1_NAME}': {delete_err}")
                     # Depending on severity, might want to stop: return False

            logger.info(f"✅ Append to {SHEET2_NAME} complete. Attempted to delete {len(row_indices_to_delete)} rows, successfully deleted {deleted_count} from '{SHEET1_NAME}'.")
            return True # Indicate rows were processed
        else:
            logger.info(f"✅ No rows found where *all* configured stores are DONE.")
            return True # No rows needed moving

    except Exception as e:
        logger.exception(f"❌ Failed during move_fully_done_to_sheet2 process: {e}")
        return False # Indicate failure